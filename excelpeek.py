#! /usr/bin/env python3
'''
ExcelPeek v0.1 - Copyright 2021 James Slaughter,
This file is part of ExcelPeek v0.1.

ExcelPeek v0.1 is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

ExcelPeek v0.1 is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with ExcelPeek v0.1.  If not, see <http://www.gnu.org/licenses/>. 
'''

#python import
import sys
import os
#import base64
#import binascii
#import struct
import re
from termcolor import colored
from openpyxl import load_workbook

#programmer generated imports
from controller import controller
from fileio import fileio

'''
Usage()
Function: Display the usage parameters when called
'''
def Usage():
    print ('Usage: [required] --file [optional] --debug --help')
    print ('Example: ./excelpeek.py --file sheet.xlsx --debug')
    print ('Required Arguments:')
    print ('--file - file being examined')
    print ('Optional Arguments:')
    print ('--list - list sheets')
    print ('--sheet - pick the sheet you want to investigate.')
    print ('--sheetstats - list the particulars of a specific sheet.')
    print ('--range - range of cells to investigate - e.g. A1:Z200.')
    print ('--dump - dump any non-empty cells')
    print ('--debug - Prints verbose logging to the screen to troubleshoot issues with a recon installation.')
    print ('--help - You\'re looking at it!')
    print ('***NOTE*** openpyxl REQUIRES that the file format be Excel 2007 .xlsx!!!')
    sys.exit(-1)
            
'''
Parse() - Parses program arguments
'''
def Parse(args):        
    option = ''

    print ('[*] Length Arguments: ' + str(len(args)))

    if (len(args) == 1):
        return -1

    print ('[*] Arguments: ')
    for i in range(len(args)):
        if args[i].startswith('--'):
            option = args[i][2:]
                
            if option == 'help':
                return -1

            if option == 'file':
                CON.file = args[i+1]
                print (option + ': ' + CON.file)

            if option == 'list':
                CON.list = True
                print (option + ': ' + str(CON.list))

            if option == 'dump':
                CON.dump = True
                print (option + ': ' + str(CON.dump))

            if option == 'sheetstats':
                CON.sheetstats = True
                print (option + ': ' + str(CON.sheetstats))

            if option == 'sheet':
                CON.sheet = args[i+1]
                print (option + ': ' + CON.sheet)

            if option == 'range':
                CON.range = args[i+1]
                print (option + ': ' + CON.range)

            if option == 'output':
                CON.output = args[i+1]
                print (option + ': ' + CON.output)
                
            if option == 'debug':
                CON.debug = True
                print (option + ': ' + str(CON.debug))               
                                     
    #These are required params so length needs to be checked after all 
    #are read through         
    if (len(CON.file) < 3):
        print ('[x] file is a required argument.')
        print ('')
        return -1

    if (CON.sheetstats == True) and (len(CON.sheet) < 3):
        print (colored('[x] sheet must be used with sheetstats.', 'red', attrs=['bold']))
        print ('')
        return -1

    if (CON.dump == True) and (len(CON.sheet) < 3):
        print (colored('[x] sheet must be used with dump.', 'red', attrs=['bold']))
        print ('')
        return -1

    if (CON.range == ''):
       CON.range = 'A1:Z200'

    print ('')   
    
    return 0

'''
List()
Function: - List all available sheets in the workbook
'''
def List():
    count = 0

    # List all the sheets in the file.
    print (colored('[*] Worksheets available are:\n', 'green', attrs=['bold']))
    for sheetname in CON.wb.sheetnames:
        print(str(count) + '. ' + sheetname)
        count +=1

    return 0

'''
sheetstats()
Function: - List all sheet stats
'''
def sheetstats():

    # Load one worksheet.
    print (colored('[*] Loading worksheet: ' + CON.sheet, 'green', attrs=['bold']))
    try:
        ws = CON.wb[CON.sheet]
    except Exception as e:
        print (colored('[x] Error: ' + str(e) + ' Terminating...', 'red', attrs=['bold']))
        return -1

    max_rows = ws.max_row
    all_rows = list(ws.rows)

    print ('[*] Max Row: ' + str(max_rows))
    print(f"[*] Found {len(all_rows)} rows of data.")

    return 0

'''
dumpsheet()
Function: - Dump all cells
'''
def dumpsheet():

    # Load one worksheet.
    print (colored('[*] Loading worksheet: ' + CON.sheet, 'green', attrs=['bold']))
    try:
        ws = CON.wb[CON.sheet]
    except Exception as e:
        print (colored('[x] Error: ' + str(e) + ' Terminating...', 'red', attrs=['bold']))
        return -1

    range_string = CON.range
    count = 0
    character = 0
    output = ''
 
    if (CON.debug == True):
        print ('[*] Length CON.output: ' + str(len(CON.output)))

    if (len(CON.output) > 3):
        filename = CON.output
    else:
        print (colored('[-] Output name not specified.  Using default: ' + CON.sheet + '.txt', 'yellow', attrs=['bold']))
        filename = CON.sheet + '.txt'
    
    FLOG = fileio()

    max_rows = ws.max_row

    col_start, col_end = re.findall("[A-Z]+", range_string)

    print (colored('[*] Output of data rows...', 'white', attrs=['bold']))
    data_rows = []
    for row in ws[range_string]:
        data_rows.append([cell.value for cell in row])

    try:
        while count < max_rows-1:
            for val in data_rows[count]:
                if ((val != None)):
                    print ('Value: ' + str(val))
                    try:
                        output += str(val) + '\r\n'
                    except ValueError as e:                   
                        print ('[x] Error - charcter: ' + str(e)) 
            FLOG.WriteLogFile(filename, output)
            output = ''
            count += 1
    except IndexError as e:                   
        print ('[x] Have reached the end of the index...')

    return 0

'''
Terminate()
Function: - Attempts to exit the program cleanly when called  
'''
     
def Terminate(exitcode):
    sys.exit(exitcode)

'''
This is the mainline section of the program and makes calls to the 
various other sections of the code
'''

if __name__ == '__main__':
    
    ret = 0

    #Stores our args
    CON = controller()
                   
    #Parses our args
    ret = Parse(sys.argv)

    #Something bad happened
    if (ret == -1):
        Usage()
        Terminate(ret)

    # Load the entire workbook.
    try:
        CON.wb = load_workbook(CON.file)
    except Exception as e:
        print (colored('[x] Error: ' + str(e) + ' Terminating...', 'red', attrs=['bold']))
        Terminate(-1)

    if (CON.list == True):
        List()
        Terminate(0)

    if (CON.sheetstats == True):
        sheetstats()

    if (CON.dump == True):
        dumpsheet()

    print ('')
    print ('[*] Program Complete')

    Terminate(0)
'''
END OF LINE
'''
